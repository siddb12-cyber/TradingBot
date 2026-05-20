"""
analytics/weekly_report.py
===========================
Weekly Decision & Trade Journal Report Generator.

Triggered every Friday at ~15:32 IST (after market close) by live_trade_tracker.py.
Reads all decision logs and trade logs for the current Mon–Fri week and produces
a richly formatted 5-sheet Excel workbook plus a Telegram summary.

Output file:  trade_logs/weekly/week_YYYY-WW.xlsx

5 Sheets:
  1. Week Overview    — headline KPIs (trades, win rate, P&L, avg score)
  2. Daily Breakdown  — one row per trading day
  3. Decision Journal — all 25-column decision rows for the week (combined)
  4. Trade Journal    — all trade log rows for the week (combined)
  5. Insights         — auto-generated text: what worked, what failed, patterns

Public entry point:
  generate_and_send(date: datetime = None) -> Path | None
"""

import logging
import os
import re
import requests
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from config.config import (
    TRADE_LOG_DIR,
    BOT_TOKEN, CHAT_ID,
)

# =========================
# LOGGING
# =========================

logger = logging.getLogger(__name__)

# =========================
# STYLE PALETTE
# =========================

_C_NAV     = "1A237E"   # Deep navy — section headers
_C_GREEN   = "1B5E20"   # Win / positive
_C_RED     = "B71C1C"   # Loss / negative
_C_AMBER   = "E65100"   # Warning
_C_BLUE    = "0D47A1"   # Neutral KPI
_C_WHITE   = "FFFFFF"
_C_LIGHT   = "F5F5F5"   # Alternating row bg
_C_PALE_G  = "E8F5E9"   # Win row bg
_C_PALE_R  = "FFEBEE"   # Loss row bg
_C_PALE_Y  = "FFFDE7"   # Neutral row bg

_THIN  = Side(style="thin",   color="BDBDBD")
_MED   = Side(style="medium", color="9E9E9E")
_BORDER_THIN = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_MED  = Border(left=_MED,  right=_MED,  top=_MED,  bottom=_MED)


def _fill(hex_c: str)  -> PatternFill: return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, color=_C_WHITE, size=10) -> Font:
    return Font(bold=bold, color=color, size=size)
def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)
def _left(wrap=False)   -> Alignment:
    return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)


def _style_header_row(ws, row: int, bg: str = _C_NAV, fg: str = _C_WHITE) -> None:
    """Apply bold header style to every cell in the given row."""
    for cell in ws[row]:
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=True, color=fg, size=10)
        cell.alignment = _center()
        cell.border    = _BORDER_THIN


def _write_kpi_card(ws, row: int, col: int, label: str, value: str,
                    bg: str = _C_NAV, fg: str = _C_WHITE) -> None:
    """Write a label+value pair into two merged cells (2 rows × 1 col)."""
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col + 1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col + 1)
    lc = ws.cell(row=row,   column=col, value=label)
    vc = ws.cell(row=row+1, column=col, value=value)
    lc.fill      = _fill(bg)
    lc.font      = _font(bold=True, color=fg, size=9)
    lc.alignment = _center()
    lc.border    = _BORDER_THIN
    vc.font      = _font(bold=True, color=bg, size=14)
    vc.alignment = _center()
    vc.border    = _BORDER_THIN


# =========================
# DATA LOADING
# =========================

def _week_date_range(ref_date: date) -> Tuple[date, date]:
    """Return (Monday, Friday) of the week containing ref_date."""
    monday = ref_date - timedelta(days=ref_date.weekday())   # weekday(): Mon=0
    friday = monday + timedelta(days=4)
    return monday, friday


def _load_decision_logs(monday: date, friday: date) -> pd.DataFrame:
    """
    Load all decisions_YYYY-MM-DD.xlsx files from trade_logs/YYYY-MM-DD/ folders
    for dates between monday and friday (inclusive).
    Returns a combined DataFrame (empty if none found).
    """
    frames: List[pd.DataFrame] = []
    current = monday
    while current <= friday:
        date_str  = current.strftime("%Y-%m-%d")
        dec_file  = TRADE_LOG_DIR / date_str / f"decisions_{date_str}.xlsx"
        if dec_file.exists():
            try:
                df = pd.read_excel(dec_file, engine="openpyxl")
                df.insert(0, "Date", date_str)   # Prepend date column
                frames.append(df)
                logger.debug("[WEEKLY] Loaded decision log: %s (%d rows)", dec_file.name, len(df))
            except Exception as exc:
                logger.warning("[WEEKLY] Could not read %s: %s", dec_file, exc)
        current += timedelta(days=1)

    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    return combined


def _load_trade_logs(monday: date, friday: date) -> pd.DataFrame:
    """
    Load all trade_log_YYYY-MM-DD.xlsx files for the week.
    Returns a combined DataFrame (empty if none found).
    """
    frames: List[pd.DataFrame] = []
    current = monday
    while current <= friday:
        date_str   = current.strftime("%Y-%m-%d")
        trade_file = TRADE_LOG_DIR / f"trade_log_{date_str}.xlsx"
        if trade_file.exists():
            try:
                df = pd.read_excel(trade_file, engine="openpyxl")
                if "Date" not in df.columns:
                    df.insert(0, "Date", date_str)
                frames.append(df)
                logger.debug("[WEEKLY] Loaded trade log: %s (%d rows)", trade_file.name, len(df))
            except Exception as exc:
                logger.warning("[WEEKLY] Could not read %s: %s", trade_file, exc)
        current += timedelta(days=1)

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# =========================
# METRICS CALCULATION
# =========================

def _calc_trade_metrics(trades: pd.DataFrame) -> dict:
    """
    Compute headline trade metrics from the trade log DataFrame.
    Handles missing columns gracefully.
    """
    m = {
        "total_trades":    0,
        "wins":            0,
        "losses":          0,
        "win_rate":        0.0,
        "total_pts":       0.0,
        "avg_score":       0.0,
        "best_trade_pts":  0.0,
        "worst_trade_pts": 0.0,
        "t1_hits":         0,
        "t2_hits":         0,
        "t3_hits":         0,
        "sl_hits":         0,
        "eod_closes":      0,
    }

    if trades.empty:
        return m

    # Only closed trades
    closed = trades[trades.get("Outcome", pd.Series(dtype=str)).notna()] if "Outcome" in trades.columns else trades

    m["total_trades"] = len(closed)
    if m["total_trades"] == 0:
        return m

    if "Points Result" in closed.columns:
        pts = pd.to_numeric(closed["Points Result"], errors="coerce").dropna()
        m["total_pts"]       = round(pts.sum(), 2)
        m["best_trade_pts"]  = round(pts.max(), 2)
        m["worst_trade_pts"] = round(pts.min(), 2)
        m["wins"]            = int((pts > 0).sum())
        m["losses"]          = int((pts <= 0).sum())

    if m["total_trades"] > 0:
        m["win_rate"] = round(m["wins"] / m["total_trades"] * 100, 1)

    if "Confidence Score" in closed.columns:
        scores = pd.to_numeric(closed["Confidence Score"], errors="coerce").dropna()
        if len(scores):
            m["avg_score"] = round(scores.mean(), 1)

    if "Outcome" in closed.columns:
        outcomes = closed["Outcome"].astype(str)
        m["t1_hits"]    = int(outcomes.str.contains("T1|TARGET_1", case=False, na=False).sum())
        m["t2_hits"]    = int(outcomes.str.contains("T2|TARGET_2", case=False, na=False).sum())
        m["t3_hits"]    = int(outcomes.str.contains("T3|TARGET_3|FULL_TARGET", case=False, na=False).sum())
        m["sl_hits"]    = int(outcomes.str.contains("SL|STOP_LOSS", case=False, na=False).sum())
        m["eod_closes"] = int(outcomes.str.contains("EOD", case=False, na=False).sum())

    return m


def _calc_daily_breakdown(trades: pd.DataFrame) -> pd.DataFrame:
    """Return a per-day summary DataFrame."""
    if trades.empty or "Date" not in trades.columns:
        return pd.DataFrame(columns=[
            "Date", "Trades", "Wins", "Losses", "Win Rate %", "Total Pts", "Avg Score"
        ])

    rows = []
    for date_val in sorted(trades["Date"].unique()):
        day = trades[trades["Date"] == date_val]
        m   = _calc_trade_metrics(day)
        rows.append({
            "Date":      date_val,
            "Trades":    m["total_trades"],
            "Wins":      m["wins"],
            "Losses":    m["losses"],
            "Win Rate %":m["win_rate"],
            "Total Pts": m["total_pts"],
            "Avg Score": m["avg_score"],
        })

    return pd.DataFrame(rows)


# =========================
# INSIGHTS GENERATOR
# =========================

def _generate_insights(
    trades: pd.DataFrame,
    decisions: pd.DataFrame,
    metrics: dict,
    monday: date,
    friday: date,
) -> List[str]:
    """
    Return a list of insight strings for the Insights sheet.
    """
    lines: List[str] = []
    week_label = f"Week {monday.strftime('%d %b')} – {friday.strftime('%d %b %Y')}"
    lines.append(f"WEEKLY INSIGHTS — {week_label}")
    lines.append("=" * 60)

    # ---- Headline summary ----
    win_emoji  = "✅" if metrics["win_rate"] >= 55 else ("⚠️" if metrics["win_rate"] >= 40 else "❌")
    lines.append(f"\n📊 PERFORMANCE SUMMARY")
    lines.append(f"  Trades executed : {metrics['total_trades']}")
    lines.append(f"  Win Rate        : {metrics['win_rate']}%  {win_emoji}")
    lines.append(f"  Net P&L (pts)   : {metrics['total_pts']:+.2f}")
    lines.append(f"  Avg Confidence  : {metrics['avg_score']:.1f}/100")
    lines.append(f"  Best trade      : +{metrics['best_trade_pts']:.2f} pts")
    lines.append(f"  Worst trade     : {metrics['worst_trade_pts']:+.2f} pts")

    # ---- Target distribution ----
    lines.append(f"\n🎯 TARGET DISTRIBUTION")
    lines.append(f"  T1 hits  : {metrics['t1_hits']}")
    lines.append(f"  T2 hits  : {metrics['t2_hits']}")
    lines.append(f"  T3 hits  : {metrics['t3_hits']}")
    lines.append(f"  SL hits  : {metrics['sl_hits']}")
    lines.append(f"  EOD close: {metrics['eod_closes']}")

    # ---- Win rate by confidence level ----
    if not trades.empty and "Confidence Level" in trades.columns and "Points Result" in trades.columns:
        lines.append(f"\n📈 WIN RATE BY CONFIDENCE LEVEL")
        for lvl in ["VERY HIGH", "HIGH", "MEDIUM", "LOW"]:
            subset = trades[trades["Confidence Level"].astype(str).str.upper() == lvl]
            if subset.empty:
                continue
            pts = pd.to_numeric(subset["Points Result"], errors="coerce").dropna()
            wins_n = int((pts > 0).sum())
            total_n = len(pts)
            wr = round(wins_n / total_n * 100, 1) if total_n else 0
            lines.append(f"  {lvl:<10}: {wins_n}/{total_n} trades = {wr}% win rate")

    # ---- Win rate by direction ----
    if not trades.empty and "Trend" in trades.columns and "Points Result" in trades.columns:
        lines.append(f"\n📉📈 WIN RATE BY DIRECTION")
        for direction in ["BULLISH", "BEARISH"]:
            subset = trades[trades["Trend"].astype(str).str.upper() == direction]
            if subset.empty:
                continue
            pts = pd.to_numeric(subset["Points Result"], errors="coerce").dropna()
            wins_n  = int((pts > 0).sum())
            total_n = len(pts)
            wr = round(wins_n / total_n * 100, 1) if total_n else 0
            avg_pts = round(pts.mean(), 2) if total_n else 0
            lines.append(f"  {direction:<8}: {wins_n}/{total_n} trades = {wr}% win rate | avg {avg_pts:+.2f} pts/trade")

    # ---- Most common blocking reasons ----
    if not decisions.empty and "Decision" in decisions.columns and "Reason" in decisions.columns:
        lines.append(f"\n🚫 MOST COMMON BLOCK REASONS")
        blocks = decisions[decisions["Decision"].isin(["BLOCKED", "LOW CONFIDENCE"])]
        if not blocks.empty and "Reason" in blocks.columns:
            reasons = blocks["Reason"].value_counts().head(5)
            for reason, count in reasons.items():
                short = str(reason)[:60]
                lines.append(f"  ({count}x) {short}")
        else:
            lines.append("  No blocks recorded this week.")

    # ---- Decision breakdown ----
    if not decisions.empty and "Decision" in decisions.columns:
        lines.append(f"\n🗂️  DECISION BREAKDOWN (all scan cycles)")
        decision_counts = decisions["Decision"].value_counts()
        total_scans = len(decisions)
        lines.append(f"  Total scan cycles: {total_scans}")
        for dec, cnt in decision_counts.items():
            pct = round(cnt / total_scans * 100, 1) if total_scans else 0
            lines.append(f"  {str(dec):<18}: {cnt} ({pct}%)")

    # ---- Score comparison: winners vs losers ----
    if not trades.empty and "Confidence Score" in trades.columns and "Points Result" in trades.columns:
        pts_col   = pd.to_numeric(trades["Points Result"],    errors="coerce")
        score_col = pd.to_numeric(trades["Confidence Score"], errors="coerce")
        winners_scores = score_col[pts_col > 0].dropna()
        losers_scores  = score_col[pts_col <= 0].dropna()
        if len(winners_scores) and len(losers_scores):
            lines.append(f"\n🔬 SCORE ANALYSIS")
            lines.append(f"  Avg score on WINNING trades: {winners_scores.mean():.1f}")
            lines.append(f"  Avg score on LOSING  trades: {losers_scores.mean():.1f}")
            diff = winners_scores.mean() - losers_scores.mean()
            if diff > 5:
                lines.append("  ✅ Higher confidence scores are correlating with wins — system calibration looks healthy.")
            elif diff < -5:
                lines.append("  ⚠️  Higher confidence trades are underperforming — review signal quality and OI/sentiment weights.")
            else:
                lines.append("  ➡️  No clear score-to-outcome correlation this week — need more data.")

    # ---- Best and worst day ----
    if not trades.empty and "Date" in trades.columns and "Points Result" in trades.columns:
        daily_pts = trades.groupby("Date")["Points Result"].apply(
            lambda x: pd.to_numeric(x, errors="coerce").sum()
        )
        if not daily_pts.empty:
            best_day  = daily_pts.idxmax()
            worst_day = daily_pts.idxmin()
            lines.append(f"\n📅 BEST / WORST DAY")
            lines.append(f"  Best  day: {best_day}  → {daily_pts[best_day]:+.2f} pts")
            lines.append(f"  Worst day: {worst_day} → {daily_pts[worst_day]:+.2f} pts")

    # ---- Recommendations ----
    lines.append(f"\n💡 RECOMMENDATIONS FOR NEXT WEEK")
    if metrics["win_rate"] >= 60:
        lines.append("  ✅ Strategy performing well. Maintain current confidence thresholds.")
    elif metrics["win_rate"] >= 45:
        lines.append("  ⚠️  Marginal performance. Review scan cycles with LOW CONFIDENCE decisions.")
        lines.append("     Consider raising CONFIDENCE_MED_THRESHOLD by 5 points.")
    else:
        lines.append("  ❌ Poor win rate. Review OI and sentiment adjustment weights.")
        lines.append("     Consider paper-running with raised CONFIDENCE_HIGH_THRESHOLD (>=75) for 1 week.")

    if metrics["sl_hits"] > metrics["wins"]:
        lines.append("  ⚠️  More SL hits than wins. Review entry timing and MTF alignment requirements.")

    if metrics["total_trades"] == 0:
        lines.append("  ℹ️  No trades executed this week. Check signal thresholds and market conditions.")

    lines.append("\n" + "=" * 60)
    lines.append("Report generated automatically by TradingBot Analytics Engine")
    lines.append(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    return lines


# =========================
# SHEET BUILDERS
# =========================

def _build_sheet_overview(wb: Workbook, metrics: dict, monday: date, friday: date) -> None:
    """Sheet 1 — Week Overview with headline KPI cards."""
    ws = wb.create_sheet("Week Overview")

    # Title
    ws.merge_cells("A1:L1")
    title = ws["A1"]
    title.value     = f"WEEKLY TRADING REPORT  |  {monday.strftime('%d %b')} – {friday.strftime('%d %b %Y')}"
    title.fill      = _fill(_C_NAV)
    title.font      = _font(bold=True, color=_C_WHITE, size=14)
    title.alignment = _center()
    title.border    = _BORDER_MED
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:L2")
    sub = ws["A2"]
    sub.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Paper Trading Mode"
    sub.fill      = _fill("3949AB")
    sub.font      = _font(bold=False, color=_C_WHITE, size=10)
    sub.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ---- KPI Section header ----
    ws.merge_cells("A4:L4")
    kpi_hdr = ws["A4"]
    kpi_hdr.value     = "📊  KEY PERFORMANCE INDICATORS"
    kpi_hdr.fill      = _fill(_C_NAV)
    kpi_hdr.font      = _font(bold=True, color=_C_WHITE, size=11)
    kpi_hdr.alignment = _left()
    kpi_hdr.border    = _BORDER_THIN
    ws.row_dimensions[4].height = 24

    # KPI cards — row 5/6 (label/value pairs)
    kpis = [
        ("Total Trades", str(metrics["total_trades"]),    _C_BLUE),
        ("Win Rate",     f"{metrics['win_rate']}%",
            _C_GREEN if metrics["win_rate"] >= 55 else (_C_AMBER if metrics["win_rate"] >= 40 else _C_RED)),
        ("Net P&L (pts)", f"{metrics['total_pts']:+.2f}",
            _C_GREEN if metrics["total_pts"] >= 0 else _C_RED),
        ("Avg Confidence", f"{metrics['avg_score']:.1f}/100", _C_BLUE),
        ("SL Hits",      str(metrics["sl_hits"]),          _C_RED),
        ("T3 Full Hits", str(metrics["t3_hits"]),          _C_GREEN),
    ]

    col_start = 1
    for label, value, colour in kpis:
        _write_kpi_card(ws, row=5, col=col_start, label=label, value=value, bg=colour)
        ws.column_dimensions[get_column_letter(col_start)].width     = 12
        ws.column_dimensions[get_column_letter(col_start+1)].width   = 4
        col_start += 2

    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 30

    # ---- Target breakdown table ----
    ws.merge_cells("A9:L9")
    ws["A9"].value     = "🎯  TARGET BREAKDOWN"
    ws["A9"].fill      = _fill(_C_NAV)
    ws["A9"].font      = _font(bold=True, color=_C_WHITE, size=11)
    ws["A9"].alignment = _left()
    ws["A9"].border    = _BORDER_THIN
    ws.row_dimensions[9].height = 24

    target_headers = ["", "T1 Partial", "T2 Partial", "T3 Full", "SL Hit", "EOD Close", "Total"]
    target_values  = [
        "Count",
        metrics["t1_hits"], metrics["t2_hits"], metrics["t3_hits"],
        metrics["sl_hits"], metrics["eod_closes"], metrics["total_trades"],
    ]

    for col_i, (hdr, val) in enumerate(zip(target_headers, target_values), start=1):
        hc = ws.cell(row=10, column=col_i, value=hdr)
        hc.fill      = _fill("37474F")
        hc.font      = _font(bold=True, color=_C_WHITE, size=10)
        hc.alignment = _center()
        hc.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(col_i)].width = 13

        vc = ws.cell(row=11, column=col_i, value=val)
        vc.font      = _font(bold=True, color="212121", size=13)
        vc.alignment = _center()
        vc.border    = _BORDER_THIN

    ws.row_dimensions[10].height = 22
    ws.row_dimensions[11].height = 30

    # Print area
    ws.print_area = "A1:L15"


def _build_sheet_daily(wb: Workbook, daily_df: pd.DataFrame) -> None:
    """Sheet 2 — Daily Breakdown."""
    ws = wb.create_sheet("Daily Breakdown")

    headers = ["Date", "Trades", "Wins", "Losses", "Win Rate %", "Total Pts", "Avg Score"]
    widths  = [14,      9,        7,       9,         12,           12,           11]

    # Header row
    for col_i, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill      = _fill(_C_NAV)
        cell.font      = _font(bold=True, color=_C_WHITE, size=10)
        cell.alignment = _center()
        cell.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    if daily_df.empty:
        ws.cell(row=2, column=1, value="No data for this week.").font = _font(color="616161")
        return

    for r_i, row in enumerate(daily_df.itertuples(index=False), start=2):
        vals = [row.Date, row.Trades, row.Wins, row.Losses,
                row._4, row._5, row._6]   # Win Rate, Total Pts, Avg Score
        bg = _C_PALE_G if row._5 > 0 else (_C_PALE_R if row._5 < 0 else _C_PALE_Y)
        for col_i, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_i, column=col_i, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(color="212121", size=10)
            cell.alignment = _center()
            cell.border    = _BORDER_THIN
        ws.row_dimensions[r_i].height = 18


def _build_sheet_decisions(wb: Workbook, decisions_df: pd.DataFrame) -> None:
    """Sheet 3 — Full Decision Journal."""
    ws = wb.create_sheet("Decision Journal")

    if decisions_df.empty:
        ws.cell(row=1, column=1, value="No decision data found for this week.").font = _font(color="616161")
        return

    # Write header
    headers = list(decisions_df.columns)
    widths  = {
        "Date": 12, "Time": 10, "Scan #": 7, "Price 5m": 11, "VWAP 5m": 11, "EMA9 5m": 11,
        "vs VWAP": 10, "vs EMA9": 10, "5m Dir": 10, "15m Dir": 10, "1h Dir": 9,
        "TF Align": 16, "Base Score": 11, "PCR": 8, "Max Pain": 10, "ATM Bias": 10,
        "OI Adj": 9, "VIX": 8, "ES/F %": 9, "Mood": 10, "Sent Adj": 9,
        "Final Score": 11, "Confidence": 14, "Decision": 18, "Reason": 48, "Trade ID": 24,
    }

    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill      = _fill(_C_NAV)
        cell.font      = _font(bold=True, color=_C_WHITE, size=9)
        cell.alignment = _center(wrap=True)
        cell.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(col_i)].width = widths.get(h, 12)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Decision colour map
    dec_bg = {
        "TRADE SIGNAL":   "1B5E20", "ACTIVE TRADE":   "BBDEFB",
        "SIDEWAYS":       "FFF9C4", "LOW CONFIDENCE": "FFE0B2",
        "BLOCKED":        "FFCDD2", "OCR ERROR":      "E1BEE7",
    }
    dec_fg = {"TRADE SIGNAL": "FFFFFF"}

    for r_i, row_vals in enumerate(dataframe_to_rows(decisions_df, index=False, header=False), start=2):
        dec = str(row_vals[headers.index("Decision")] if "Decision" in headers else "")
        bg  = dec_bg.get(dec, "FFFFFF" if r_i % 2 == 0 else _C_LIGHT)
        fg  = dec_fg.get(dec, "212121")
        for col_i, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=r_i, column=col_i, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(color=fg, size=9)
            cell.alignment = _center()
            cell.border    = _BORDER_THIN
        ws.row_dimensions[r_i].height = 16


def _build_sheet_trades(wb: Workbook, trades_df: pd.DataFrame) -> None:
    """Sheet 4 — Trade Journal."""
    ws = wb.create_sheet("Trade Journal")

    if trades_df.empty:
        ws.cell(row=1, column=1, value="No trades found for this week.").font = _font(color="616161")
        return

    headers = list(trades_df.columns)
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill      = _fill(_C_NAV)
        cell.font      = _font(bold=True, color=_C_WHITE, size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(col_i)].width = max(12, len(str(h)) + 2)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    pts_col_idx = headers.index("Points Result") if "Points Result" in headers else None

    for r_i, row_vals in enumerate(dataframe_to_rows(trades_df, index=False, header=False), start=2):
        pts = None
        if pts_col_idx is not None:
            try:
                pts = float(row_vals[pts_col_idx])
            except (TypeError, ValueError):
                pts = None

        bg = (_C_PALE_G if pts and pts > 0 else
              _C_PALE_R if pts and pts < 0 else
              _C_PALE_Y)

        for col_i, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=r_i, column=col_i, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(color="212121", size=10)
            cell.alignment = _center()
            cell.border    = _BORDER_THIN
        ws.row_dimensions[r_i].height = 18


def _build_sheet_insights(wb: Workbook, insight_lines: List[str]) -> None:
    """Sheet 5 — Insights (text-based analysis)."""
    ws = wb.create_sheet("Insights")

    ws.merge_cells("A1:C1")
    ws["A1"].value     = "📋  WEEKLY INSIGHTS & ANALYSIS"
    ws["A1"].fill      = _fill(_C_NAV)
    ws["A1"].font      = _font(bold=True, color=_C_WHITE, size=13)
    ws["A1"].alignment = _left()
    ws["A1"].border    = _BORDER_THIN
    ws.row_dimensions[1].height = 30

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 20

    for r_i, line in enumerate(insight_lines, start=2):
        is_section = line.startswith(("=", "📊", "🎯", "📈", "📉", "🚫", "🗂️", "🔬", "📅", "💡", "WEEKLY"))
        cell = ws.cell(row=r_i, column=2, value=line)

        if is_section or line.startswith("==="):
            cell.font      = _font(bold=True, color=_C_NAV, size=10)
            cell.fill      = _fill("E8EAF6")
        else:
            cell.font      = Font(color="212121", size=10, name="Calibri")

        cell.alignment = _left(wrap=True)
        cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))
        ws.row_dimensions[r_i].height = 16


# =========================
# TELEGRAM SUMMARY
# =========================

def _send_weekly_telegram(metrics: dict, monday: date, friday: date) -> None:
    """Send a weekly summary message to Telegram."""
    if not BOT_TOKEN or not CHAT_ID:
        logger.warning("[WEEKLY] Telegram credentials missing — skipping Telegram alert.")
        return

    week_str = f"{monday.strftime('%d %b')} – {friday.strftime('%d %b %Y')}"
    sep      = "--" * 16

    win_flag  = "✅" if metrics["win_rate"] >= 55 else ("⚠️" if metrics["win_rate"] >= 40 else "❌")
    pnl_flag  = "📈" if metrics["total_pts"] >= 0 else "📉"

    msg = (
        f"WEEKLY TRADING REPORT\n{sep}\n"
        f"Period      : {week_str}\n"
        f"{sep}\n"
        f"Total Trades: {metrics['total_trades']}\n"
        f"Win Rate    : {metrics['win_rate']}%  {win_flag}\n"
        f"Net P&L     : {metrics['total_pts']:+.2f} pts  {pnl_flag}\n"
        f"Avg Score   : {metrics['avg_score']:.1f}/100\n"
        f"{sep}\n"
        f"T1 Hits  : {metrics['t1_hits']}\n"
        f"T2 Hits  : {metrics['t2_hits']}\n"
        f"T3 Hits  : {metrics['t3_hits']}\n"
        f"SL Hits  : {metrics['sl_hits']}\n"
        f"EOD Close: {metrics['eod_closes']}\n"
        f"{sep}\n"
        f"Best Trade : +{metrics['best_trade_pts']:.2f} pts\n"
        f"Worst Trade: {metrics['worst_trade_pts']:+.2f} pts\n"
        f"{sep}\n"
        f"Full weekly journal saved to trade_logs/weekly/\n"
        f"[Paper Trading Mode]"
    )

    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        r   = requests.post(url, data={"chat_id": CHAT_ID, "text": msg}, timeout=10)
        if r.status_code == 200:
            logger.info("[WEEKLY] Telegram weekly summary sent.")
        else:
            logger.warning("[WEEKLY] Telegram send failed: %s %s", r.status_code, r.text[:100])
    except Exception as exc:
        logger.error("[WEEKLY] Telegram error: %s", exc)


# =========================
# MAIN ENTRY POINT
# =========================

def generate_and_send(ref_date: Optional[datetime] = None) -> Optional[Path]:
    """
    Generate the weekly report Excel file and send Telegram summary.

    Parameters
    ----------
    ref_date : Any date within the target week (defaults to today).

    Returns
    -------
    Path to the generated .xlsx file, or None on failure.
    """
    ref = (ref_date or datetime.now()).date()
    monday, friday = _week_date_range(ref)
    week_num = monday.isocalendar()[1]

    logger.info(
        "[WEEKLY] Generating report for week %d: %s → %s",
        week_num, monday, friday
    )

    # ---- Load data ----
    decisions_df = _load_decision_logs(monday, friday)
    trades_df    = _load_trade_logs(monday, friday)

    logger.info(
        "[WEEKLY] Data loaded — decisions: %d rows | trades: %d rows",
        len(decisions_df), len(trades_df)
    )

    # ---- Calculate metrics ----
    metrics   = _calc_trade_metrics(trades_df)
    daily_df  = _calc_daily_breakdown(trades_df)
    insights  = _generate_insights(trades_df, decisions_df, metrics, monday, friday)

    # ---- Build workbook ----
    wb = Workbook()
    wb.remove(wb.active)   # Remove default empty sheet

    _build_sheet_overview(wb, metrics, monday, friday)
    _build_sheet_daily(wb, daily_df)
    _build_sheet_decisions(wb, decisions_df)
    _build_sheet_trades(wb, trades_df)
    _build_sheet_insights(wb, insights)

    # ---- Save ----
    weekly_dir = TRADE_LOG_DIR / "weekly"
    weekly_dir.mkdir(parents=True, exist_ok=True)

    out_file = weekly_dir / f"week_{monday.strftime('%Y')}-W{week_num:02d}.xlsx"
    tmp_file = out_file.with_suffix(".tmp.xlsx")

    try:
        wb.save(str(tmp_file))
        os.replace(str(tmp_file), str(out_file))
        logger.info("[WEEKLY] Report saved → %s", out_file.name)
    except Exception as exc:
        logger.error("[WEEKLY] Failed to save report: %s", exc)
        try:
            if tmp_file.exists():
                tmp_file.unlink()
        except Exception:
            pass
        return None

    # ---- Telegram ----
    _send_weekly_telegram(metrics, monday, friday)

    return out_file
