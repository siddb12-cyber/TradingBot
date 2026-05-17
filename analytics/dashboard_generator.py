"""
analytics/dashboard_generator.py
==================================
Generates a professional Dark-Pro Bloomberg-style trading dashboard
as a self-contained HTML file deployable to GitHub Pages.

Reads:
  - trade_logs/trade_log_YYYY-MM-DD.xlsx  (all dates)
  - data/paper_validation_metrics.json    (readiness + validation)
  - data/trade_state.json                 (active trade)

Writes:
  - docs/index.html  (GitHub Pages compatible)

Usage:
    python -m analytics.dashboard_generator        # generate only
    python -m analytics.dashboard_generator --open  # generate + open in browser

Schedule:
    Called automatically by analytics/paper_trading_report.py at EOD.
    Also called by deploy_dashboard.bat for manual push to GitHub Pages.
"""

import json
import logging
import os
import sys
import webbrowser
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
from config.config import (
    BASE_DIR,
    TRADE_LOG_DIR,
    DATA_DIR,
    PAPER_TRADING_MODE,
    PAPER_TRADING_VALIDATION_END,
)

logger = logging.getLogger(__name__)
DOCS_DIR   = BASE_DIR / "docs"
OUTPUT_HTML = DOCS_DIR / "index.html"
STATE_FILE  = DATA_DIR / "trade_state.json"
METRICS_FILE = DATA_DIR / "paper_validation_metrics.json"

# =========================
# DATA LOADING
# =========================

def _load_xlsx_files() -> list:
    """Read all trade_log_*.xlsx files and return list of row dicts."""
    rows = []
    for xlsx in sorted(TRADE_LOG_DIR.glob("trade_log_*.xlsx")):
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append(dict(zip(headers, row)))
        except Exception as exc:
            logger.warning(f"[DASH] Could not read {xlsx.name}: {exc}")
    return rows


def _load_validation_metrics() -> dict:
    if METRICS_FILE.exists():
        try:
            return json.loads(METRICS_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _load_active_trade() -> Optional[dict]:
    if STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text(encoding="utf-8"))
            if state.get("status") == "OPEN":
                return state
        except Exception:
            pass
    return None


# =========================
# KPI COMPUTATION
# =========================

def _safe_float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


def compute_kpis(rows: list) -> dict:
    """Compute all portfolio KPIs from trade rows."""
    closed = [r for r in rows if r.get("Outcome") and r.get("Outcome") != "OPEN"]
    open_trades = [r for r in rows if r.get("Trade Status") == "OPEN" or r.get("Outcome") is None]

    total = len(closed)
    wins  = [r for r in closed if str(r.get("Outcome","")).startswith("T")]
    sl_hits = [r for r in closed if str(r.get("Outcome","")).startswith("SL")]
    win_rate = round(len(wins) / total * 100, 1) if total else 0

    points = [_safe_float(r.get("Points Result")) for r in closed if r.get("Points Result") is not None]
    total_pts  = round(sum(points), 1)
    best_pts   = round(max(points), 1) if points else 0
    worst_pts  = round(min(points), 1) if points else 0

    # Cumulative P&L series for chart (per trade)
    cumulative = []
    running = 0
    for r in rows:
        pts = _safe_float(r.get("Points Result"))
        running = round(running + pts, 1)
        cumulative.append({
            "date":  str(r.get("Date", "")),
            "time":  str(r.get("Time", "")),
            "cum":   running,
            "pts":   pts,
        })

    # Daily P&L
    daily_pnl: dict = {}
    for r in rows:
        d = str(r.get("Date", "unknown"))
        pts = _safe_float(r.get("Points Result"))
        daily_pnl[d] = round(daily_pnl.get(d, 0) + pts, 1)

    # Confidence
    conf_scores = [_safe_float(r.get("Confidence Score")) for r in rows if r.get("Confidence Score")]
    avg_conf = round(sum(conf_scores) / len(conf_scores), 1) if conf_scores else 0

    # Max drawdown
    max_dd = 0
    peak = 0
    running2 = 0
    for p in points:
        running2 += p
        if running2 > peak:
            peak = running2
        dd = peak - running2
        if dd > max_dd:
            max_dd = dd
    max_dd = round(max_dd, 1)

    # Streak
    streak = 0
    streak_type = ""
    for r in reversed(closed):
        outcome = str(r.get("Outcome", ""))
        if not streak_type:
            streak_type = "W" if outcome.startswith("T") else "L"
        if streak_type == "W" and outcome.startswith("T"):
            streak += 1
        elif streak_type == "L" and outcome.startswith("SL"):
            streak += 1
        else:
            break

    # Lots stats
    lots = [_safe_float(r.get("Lots", 1)) for r in rows if r.get("Lots")]
    avg_lots = round(sum(lots)/len(lots), 1) if lots else 1

    return {
        "total_trades":   total,
        "open_count":     len(open_trades),
        "win_count":      len(wins),
        "sl_count":       len(sl_hits),
        "win_rate":       win_rate,
        "total_pts":      total_pts,
        "best_pts":       best_pts,
        "worst_pts":      worst_pts,
        "avg_conf":       avg_conf,
        "max_drawdown":   max_dd,
        "streak":         streak,
        "streak_type":    streak_type,
        "avg_lots":       avg_lots,
        "cumulative":     cumulative,
        "daily_pnl":      [{"date": k, "pnl": v} for k, v in sorted(daily_pnl.items())],
        "trade_rows":     [_serialize_row(r) for r in rows],
    }


def _serialize_row(r: dict) -> dict:
    """Prepare a single trade row for JSON embedding."""
    return {
        "date":       str(r.get("Date", "")),
        "time":       str(r.get("Time", "")),
        "signal":     str(r.get("Trade Signal", "")),
        "trend":      str(r.get("Trend", "")),
        "price":      _safe_float(r.get("Current Price")),
        "vwap":       _safe_float(r.get("VWAP")),
        "outcome":    str(r.get("Outcome") or "OPEN"),
        "points":     _safe_float(r.get("Points Result")),
        "confidence": _safe_float(r.get("Confidence Score")),
        "conf_level": str(r.get("Confidence Level") or ""),
        "lots":       _safe_float(r.get("Lots", 1)),
        "mtf":        str(r.get("MTF Alignment") or ""),
        "tf5":        str(r.get("TF 5m") or ""),
        "tf15":       str(r.get("TF 15m") or ""),
        "tf1h":       str(r.get("TF 1h") or ""),
        "trade_id":   str(r.get("Trade ID") or ""),
        "status":     str(r.get("Trade Status") or ""),
    }


# =========================
# HTML GENERATOR
# =========================

def build_html(kpis: dict, validation: dict, active_trade: Optional[dict]) -> str:
    updated_at = datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")
    paper_badge = "PAPER MODE" if PAPER_TRADING_MODE else "LIVE MODE"
    paper_color = "#f59e0b" if PAPER_TRADING_MODE else "#ff3366"
    val_end     = PAPER_TRADING_VALIDATION_END

    # Readiness from validation metrics
    readiness_score  = 0
    readiness_status = "NOT READY"
    val_metrics      = {}
    if validation:
        cumulative_data = validation.get("cumulative") or validation
        readiness_score  = cumulative_data.get("readiness_score", 0)
        readiness_status = cumulative_data.get("readiness_status", "NOT READY")
        val_metrics = {
            "signal_accuracy":    round(cumulative_data.get("signal_accuracy", 0) * 100, 1),
            "sl_ratio":           round(cumulative_data.get("sl_ratio", 0) * 100, 1),
            "avg_confidence":     round(cumulative_data.get("avg_confidence_score", 0), 1),
            "max_consec_losses":  cumulative_data.get("max_consecutive_losses", 0),
            "trade_count":        cumulative_data.get("total_trades", kpis["total_trades"]),
            "rejection_count":    cumulative_data.get("rejection_count", 0),
        }
    else:
        val_metrics = {
            "signal_accuracy":   kpis["win_rate"],
            "sl_ratio":          round(kpis["sl_count"] / max(kpis["total_trades"],1)*100, 1),
            "avg_confidence":    kpis["avg_conf"],
            "max_consec_losses": 0,
            "trade_count":       kpis["total_trades"],
            "rejection_count":   0,
        }

    readiness_color = {
        "READY":       "#00d084",
        "CONDITIONAL": "#f59e0b",
        "NOT READY":   "#ff3366",
    }.get(readiness_status, "#ff3366")

    active_html = ""
    if active_trade:
        active_html = f"""
        <div class="active-trade-banner">
          <span class="live-dot"></span>
          <strong>ACTIVE TRADE:</strong> &nbsp;
          {active_trade.get('trade_signal','')} &nbsp;|&nbsp;
          Entry: <strong>{active_trade.get('entry_price','')}</strong> &nbsp;|&nbsp;
          SL: <strong>{active_trade.get('stop_loss','')}</strong> &nbsp;|&nbsp;
          ID: <code>{active_trade.get('trade_id','')}</code>
        </div>"""

    kpis_json      = json.dumps(kpis, default=str)
    val_json       = json.dumps(val_metrics, default=str)

    total_pts_color = "#00d084" if kpis["total_pts"] >= 0 else "#ff3366"
    best_color      = "#00d084" if kpis["best_pts"] >= 0 else "#ff3366"
    worst_color     = "#ff3366" if kpis["worst_pts"] < 0 else "#00d084"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>TradingBot Portfolio Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet"/>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  :root{{
    --bg:#070d1a;--bg-card:#0c1526;--bg-card2:#0f1c30;--border:#1a2d4a;
    --green:#00d084;--red:#ff3366;--blue:#3b82f6;--yellow:#f59e0b;--purple:#8b5cf6;
    --text:#e2e8f0;--text2:#94a3b8;--text3:#475569;
    --font:'Inter',sans-serif;--mono:'JetBrains Mono',monospace;
  }}
  html,body{{background:var(--bg);color:var(--text);font-family:var(--font);min-height:100vh;font-size:14px}}
  .app{{max-width:1400px;margin:0 auto;padding:20px 24px}}

  /* HEADER */
  .header{{display:flex;align-items:center;justify-content:space-between;padding:16px 0 24px;border-bottom:1px solid var(--border);margin-bottom:24px}}
  .header-left{{display:flex;align-items:center;gap:14px}}
  .header-title{{font-size:20px;font-weight:700;letter-spacing:0.5px}}
  .header-sub{{font-size:12px;color:var(--text2);margin-top:2px}}
  .badge{{padding:4px 10px;border-radius:4px;font-size:11px;font-weight:600;letter-spacing:0.8px}}
  .badge-paper{{background:rgba(245,158,11,0.15);color:#f59e0b;border:1px solid rgba(245,158,11,0.3)}}
  .badge-live{{background:rgba(255,51,102,0.15);color:#ff3366;border:1px solid rgba(255,51,102,0.3)}}
  .header-right{{text-align:right;font-size:12px;color:var(--text2)}}
  .update-time{{font-family:var(--mono);font-size:11px;color:var(--text3)}}

  /* ACTIVE TRADE BANNER */
  .active-trade-banner{{
    background:rgba(59,130,246,0.1);border:1px solid rgba(59,130,246,0.3);
    border-radius:8px;padding:10px 16px;margin-bottom:20px;
    display:flex;align-items:center;gap:10px;font-size:13px;color:#93c5fd
  }}
  .live-dot{{width:8px;height:8px;border-radius:50%;background:#3b82f6;
    box-shadow:0 0 8px #3b82f6;animation:pulse 1.5s infinite}}
  @keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:0.4}}}}

  /* KPI CARDS */
  .kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:14px;margin-bottom:24px}}
  .kpi-card{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:18px 16px;position:relative;overflow:hidden}}
  .kpi-card::before{{content:'';position:absolute;top:0;left:0;right:0;height:2px}}
  .kpi-green::before{{background:linear-gradient(90deg,var(--green),transparent)}}
  .kpi-red::before{{background:linear-gradient(90deg,var(--red),transparent)}}
  .kpi-blue::before{{background:linear-gradient(90deg,var(--blue),transparent)}}
  .kpi-yellow::before{{background:linear-gradient(90deg,var(--yellow),transparent)}}
  .kpi-purple::before{{background:linear-gradient(90deg,var(--purple),transparent)}}
  .kpi-label{{font-size:10px;font-weight:500;letter-spacing:1px;color:var(--text3);text-transform:uppercase;margin-bottom:8px}}
  .kpi-value{{font-size:26px;font-weight:700;font-family:var(--mono);line-height:1}}
  .kpi-sub{{font-size:11px;color:var(--text2);margin-top:6px}}

  /* CHARTS */
  .charts-grid{{display:grid;grid-template-columns:2fr 1fr;gap:16px;margin-bottom:24px}}
  .card{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:20px}}
  .card-title{{font-size:12px;font-weight:600;letter-spacing:1px;color:var(--text2);text-transform:uppercase;margin-bottom:16px;display:flex;align-items:center;gap:8px}}
  .card-title span{{width:3px;height:14px;border-radius:2px;display:inline-block}}
  .chart-wrap{{position:relative;height:220px}}

  /* VALIDATION PANEL */
  .val-grid{{display:grid;grid-template-columns:200px 1fr;gap:16px;margin-bottom:24px;align-items:start}}
  .readiness-card{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:20px;text-align:center}}
  .readiness-score{{font-size:56px;font-weight:700;font-family:var(--mono);line-height:1;margin:12px 0 8px}}
  .readiness-label{{font-size:11px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;padding:4px 12px;border-radius:4px;display:inline-block;margin-bottom:8px}}
  .readiness-sub{{font-size:11px;color:var(--text3)}}
  .metrics-card{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:20px}}
  .metric-row{{display:flex;align-items:center;gap:12px;margin-bottom:14px}}
  .metric-row:last-child{{margin-bottom:0}}
  .metric-name{{font-size:11px;color:var(--text2);width:130px;flex-shrink:0;letter-spacing:0.3px}}
  .metric-bar-wrap{{flex:1;background:#0a1628;border-radius:3px;height:6px;overflow:hidden}}
  .metric-bar{{height:100%;border-radius:3px;transition:width 0.6s ease}}
  .metric-val{{font-size:12px;font-family:var(--mono);font-weight:600;width:48px;text-align:right;flex-shrink:0}}

  /* TRADE TABLE */
  .table-card{{background:var(--bg-card);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:24px}}
  .table-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}}
  .filter-btns{{display:flex;gap:6px}}
  .filter-btn{{padding:5px 12px;border-radius:5px;border:1px solid var(--border);background:transparent;color:var(--text2);font-size:11px;cursor:pointer;font-family:var(--font);transition:all 0.2s}}
  .filter-btn:hover,.filter-btn.active{{background:var(--blue);border-color:var(--blue);color:#fff}}
  .table-wrap{{overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:12px}}
  th{{text-align:left;padding:8px 12px;color:var(--text3);font-size:10px;letter-spacing:0.8px;text-transform:uppercase;border-bottom:1px solid var(--border);font-weight:500}}
  td{{padding:10px 12px;border-bottom:1px solid rgba(26,45,74,0.5);vertical-align:middle}}
  tr:hover td{{background:rgba(59,130,246,0.04)}}
  tr:last-child td{{border-bottom:none}}
  .signal-cell{{font-family:var(--mono);font-size:11px;font-weight:500}}
  .outcome-badge{{padding:3px 9px;border-radius:4px;font-size:10px;font-weight:600;letter-spacing:0.5px;font-family:var(--mono)}}
  .outcome-T1,.outcome-T2,.outcome-T3{{background:rgba(0,208,132,0.12);color:var(--green)}}
  .outcome-SL{{background:rgba(255,51,102,0.12);color:var(--red)}}
  .outcome-EOD{{background:rgba(148,163,184,0.12);color:var(--text2)}}
  .outcome-OPEN{{background:rgba(59,130,246,0.12);color:var(--blue)}}
  .conf-HIGH{{color:var(--green)}}
  .conf-MEDIUM{{color:var(--yellow)}}
  .conf-LOW{{color:var(--red)}}
  .pts-pos{{color:var(--green);font-family:var(--mono);font-weight:600}}
  .pts-neg{{color:var(--red);font-family:var(--mono);font-weight:600}}
  .pts-zero{{color:var(--text2);font-family:var(--mono)}}
  .empty-state{{text-align:center;padding:40px;color:var(--text3)}}
  .empty-icon{{font-size:32px;margin-bottom:8px}}

  /* FOOTER */
  .footer{{text-align:center;padding:16px 0;color:var(--text3);font-size:11px;border-top:1px solid var(--border);margin-top:8px}}

  @media(max-width:900px){{
    .charts-grid{{grid-template-columns:1fr}}
    .val-grid{{grid-template-columns:1fr}}
    .kpi-grid{{grid-template-columns:repeat(2,1fr)}}
  }}
</style>
</head>
<body>
<div class="app">

  <!-- HEADER -->
  <div class="header">
    <div class="header-left">
      <div>
        <div class="header-title">⚡ TradingBot Portfolio</div>
        <div class="header-sub">NIFTY Options &nbsp;·&nbsp; Intraday &nbsp;·&nbsp; AI-Assisted</div>
      </div>
      <span class="badge badge-paper" style="color:{paper_color};border-color:rgba(245,158,11,0.3)">{paper_badge}</span>
    </div>
    <div class="header-right">
      <div style="font-size:12px;color:var(--text2)">Managed by TradingBot</div>
      <div class="update-time">Updated: {updated_at}</div>
      <div style="font-size:11px;color:var(--text3);margin-top:2px">Validation until {val_end}</div>
    </div>
  </div>

  {active_html}

  <!-- KPI CARDS -->
  <div class="kpi-grid">
    <div class="kpi-card kpi-{'green' if kpis['total_pts']>=0 else 'red'}">
      <div class="kpi-label">Total P&amp;L</div>
      <div class="kpi-value" style="color:{total_pts_color}">{'+' if kpis['total_pts']>=0 else ''}{kpis['total_pts']}</div>
      <div class="kpi-sub">Points &nbsp;|&nbsp; Paper Mode</div>
    </div>
    <div class="kpi-card kpi-blue">
      <div class="kpi-label">Win Rate</div>
      <div class="kpi-value" style="color:var(--blue)">{kpis['win_rate']}%</div>
      <div class="kpi-sub">{kpis['win_count']}W / {kpis['sl_count']}L of {kpis['total_trades']} trades</div>
    </div>
    <div class="kpi-card kpi-purple">
      <div class="kpi-label">Avg Confidence</div>
      <div class="kpi-value" style="color:var(--purple)">{kpis['avg_conf']}</div>
      <div class="kpi-sub">out of 100 &nbsp;|&nbsp; MTF Score</div>
    </div>
    <div class="kpi-card kpi-green">
      <div class="kpi-label">Best Trade</div>
      <div class="kpi-value" style="color:{best_color}">+{kpis['best_pts']}</div>
      <div class="kpi-sub">Points</div>
    </div>
    <div class="kpi-card kpi-red">
      <div class="kpi-label">Worst Trade</div>
      <div class="kpi-value" style="color:{worst_color}">{kpis['worst_pts']}</div>
      <div class="kpi-sub">Points</div>
    </div>
    <div class="kpi-card kpi-yellow">
      <div class="kpi-label">Max Drawdown</div>
      <div class="kpi-value" style="color:var(--yellow)">{kpis['max_drawdown']}</div>
      <div class="kpi-sub">Points peak-to-trough</div>
    </div>
    <div class="kpi-card kpi-{'green' if kpis['streak_type']=='W' else 'red' if kpis['streak_type']=='L' else 'blue'}">
      <div class="kpi-label">Current Streak</div>
      <div class="kpi-value" style="color:{'var(--green)' if kpis['streak_type']=='W' else 'var(--red)' if kpis['streak_type']=='L' else 'var(--blue)'}">
        {kpis['streak']}{kpis['streak_type'] or '–'}
      </div>
      <div class="kpi-sub">{'Winning' if kpis['streak_type']=='W' else 'Losing' if kpis['streak_type']=='L' else 'No trades yet'} streak</div>
    </div>
    <div class="kpi-card kpi-blue">
      <div class="kpi-label">Open Trades</div>
      <div class="kpi-value" style="color:var(--blue)">{kpis['open_count']}</div>
      <div class="kpi-sub">Active positions</div>
    </div>
  </div>

  <!-- CHARTS -->
  <div class="charts-grid">
    <div class="card">
      <div class="card-title"><span style="background:var(--green)"></span>Cumulative P&amp;L</div>
      <div class="chart-wrap"><canvas id="cumChart"></canvas></div>
    </div>
    <div class="card">
      <div class="card-title"><span style="background:var(--blue)"></span>Daily P&amp;L</div>
      <div class="chart-wrap"><canvas id="dailyChart"></canvas></div>
    </div>
  </div>

  <!-- VALIDATION PANEL -->
  <div class="val-grid">
    <div class="readiness-card">
      <div class="card-title" style="justify-content:center"><span style="background:var(--purple)"></span>Readiness</div>
      <div class="readiness-score" style="color:{readiness_color}">{readiness_score}</div>
      <div class="readiness-label" style="background:rgba(0,0,0,0.3);color:{readiness_color};border:1px solid {readiness_color}30">{readiness_status}</div>
      <div class="readiness-sub" style="margin-top:10px">out of 100 points</div>
      <div class="readiness-sub" style="margin-top:4px">Live threshold: 75</div>
    </div>
    <div class="metrics-card">
      <div class="card-title"><span style="background:var(--purple)"></span>Validation Metrics</div>
      <div id="metricsContainer"></div>
    </div>
  </div>

  <!-- TRADE HISTORY -->
  <div class="table-card">
    <div class="table-header">
      <div class="card-title" style="margin-bottom:0"><span style="background:var(--blue)"></span>Trade History</div>
      <div class="filter-btns">
        <button class="filter-btn active" onclick="filterTrades('ALL')">All</button>
        <button class="filter-btn" onclick="filterTrades('WIN')">Wins</button>
        <button class="filter-btn" onclick="filterTrades('LOSS')">Losses</button>
        <button class="filter-btn" onclick="filterTrades('OPEN')">Open</button>
      </div>
    </div>
    <div class="table-wrap">
      <table id="tradeTable">
        <thead><tr>
          <th>Date</th><th>Time</th><th>Signal</th><th>Entry</th>
          <th>Outcome</th><th>Points</th><th>Confidence</th><th>MTF</th><th>TF 5m</th>
        </tr></thead>
        <tbody id="tradeBody"></tbody>
      </table>
    </div>
  </div>

  <div class="footer">TradingBot &nbsp;·&nbsp; Haus and Kinder / Rivermoor &nbsp;·&nbsp; Samara Retail India Pvt Ltd &nbsp;·&nbsp; Paper Trading Active</div>
</div>

<script>
const KPIS = {kpis_json};
const VAL  = {val_json};

// Chart defaults
Chart.defaults.color = '#475569';
Chart.defaults.borderColor = '#1a2d4a';
Chart.defaults.font.family = "'JetBrains Mono', monospace";
Chart.defaults.font.size = 11;

// ── Cumulative P&L Chart ──
const cumData = KPIS.cumulative || [];
const cumLabels = cumData.map((_,i)=>i+1);
const cumValues = cumData.map(d=>d.cum);
new Chart(document.getElementById('cumChart'), {{
  type: 'line',
  data: {{
    labels: cumLabels,
    datasets: [{{
      label: 'Cumulative P&L (pts)',
      data: cumValues,
      borderColor: cumValues.length && cumValues[cumValues.length-1] >= 0 ? '#00d084' : '#ff3366',
      backgroundColor: function(ctx){{
        const g = ctx.chart.ctx.createLinearGradient(0,0,0,200);
        g.addColorStop(0, cumValues.length && cumValues[cumValues.length-1]>=0 ? 'rgba(0,208,132,0.15)' : 'rgba(255,51,102,0.15)');
        g.addColorStop(1, 'rgba(0,0,0,0)');
        return g;
      }},
      borderWidth: 2,
      fill: true,
      tension: 0.3,
      pointRadius: cumValues.length > 30 ? 0 : 3,
      pointHoverRadius: 5,
    }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display: false }}, tooltip: {{
      callbacks: {{ label: ctx => (ctx.parsed.y>=0?'+':'')+ctx.parsed.y+' pts' }}
    }}}},
    scales: {{
      x: {{ grid: {{ display:false }}, ticks: {{ maxTicksLimit:8 }} }},
      y: {{ grid: {{ color:'#0f1e30' }}, ticks: {{ callback: v=>(v>=0?'+':'')+v }} }}
    }}
  }}
}});

// ── Daily P&L Chart ──
const dailyData = KPIS.daily_pnl || [];
const dLabels = dailyData.map(d=>d.date);
const dValues = dailyData.map(d=>d.pnl);
new Chart(document.getElementById('dailyChart'), {{
  type: 'bar',
  data: {{
    labels: dLabels,
    datasets: [{{
      label: 'Daily P&L (pts)',
      data: dValues,
      backgroundColor: dValues.map(v=>v>=0?'rgba(0,208,132,0.7)':'rgba(255,51,102,0.7)'),
      borderColor: dValues.map(v=>v>=0?'#00d084':'#ff3366'),
      borderWidth: 1, borderRadius: 3,
    }}]
  }},
  options: {{
    responsive: true, maintainAspectRatio: false,
    plugins: {{ legend: {{ display:false }}, tooltip: {{
      callbacks: {{ label: ctx => (ctx.parsed.y>=0?'+':'')+ctx.parsed.y+' pts' }}
    }}}},
    scales: {{
      x: {{ grid: {{ display:false }}, ticks: {{ maxTicksLimit:6 }} }},
      y: {{ grid: {{ color:'#0f1e30' }} }}
    }}
  }}
}});

// ── Validation Metrics ──
const metricDefs = [
  {{ key:'signal_accuracy', label:'Signal Accuracy', unit:'%', target:55, invert:false, color:'#00d084' }},
  {{ key:'sl_ratio',        label:'SL Hit Ratio',    unit:'%', target:40, invert:true,  color:'#ff3366' }},
  {{ key:'avg_confidence',  label:'Avg Confidence',  unit:'',  target:60, invert:false, color:'#8b5cf6' }},
  {{ key:'max_consec_losses',label:'Max Consec Loss',unit:'',  target:3,  invert:true,  color:'#f59e0b' }},
  {{ key:'trade_count',     label:'Trades Logged',   unit:'',  target:20, invert:false, color:'#3b82f6' }},
  {{ key:'rejection_count', label:'Rejections',      unit:'',  target:100,invert:false, color:'#475569' }},
];
const mc = document.getElementById('metricsContainer');
metricDefs.forEach(m => {{
  const val = VAL[m.key] || 0;
  const pct = Math.min((val / (m.target || 100)) * 100, 100);
  const isGood = m.invert ? val <= m.target : val >= m.target;
  const barColor = isGood ? m.color : (m.invert ? '#00d084' : '#ff3366');
  mc.innerHTML += `
    <div class="metric-row">
      <div class="metric-name">${{m.label}}</div>
      <div class="metric-bar-wrap">
        <div class="metric-bar" style="width:${{pct}}%;background:${{barColor}}"></div>
      </div>
      <div class="metric-val" style="color:${{barColor}}">${{val}}${{m.unit}}</div>
    </div>`;
}});

// ── Trade Table ──
const trades = KPIS.trade_rows || [];
function renderTable(filter) {{
  document.querySelectorAll('.filter-btn').forEach(b=>b.classList.toggle('active', b.textContent===filter));
  const tbody = document.getElementById('tradeBody');
  const filtered = trades.filter(t=>{{
    if(filter==='ALL')   return true;
    if(filter==='WIN')   return t.outcome.startsWith('T');
    if(filter==='LOSS')  return t.outcome.startsWith('SL');
    if(filter==='OPEN')  return t.outcome==='OPEN'||t.status==='OPEN';
    return true;
  }});
  if(!filtered.length){{
    tbody.innerHTML='<tr><td colspan="9"><div class="empty-state"><div class="empty-icon">📊</div><div>No trades yet — system will populate data as trades execute</div></div></td></tr>';
    return;
  }}
  tbody.innerHTML = [...filtered].reverse().map(t=>{{
    const oc = t.outcome.startsWith('T')?'T1':'outcome-'+t.outcome.split('_')[0];
    const ptsCls = t.points>0?'pts-pos':t.points<0?'pts-neg':'pts-zero';
    const ptsStr = t.points!==0?(t.points>0?'+':'')+t.points+'pts':'–';
    return `<tr>
      <td style="color:var(--text2);font-family:var(--mono)">${{t.date}}</td>
      <td style="color:var(--text3);font-family:var(--mono)">${{t.time}}</td>
      <td class="signal-cell">${{t.signal}}</td>
      <td style="font-family:var(--mono)">${{t.price||'–'}}</td>
      <td><span class="outcome-badge outcome-${{t.outcome.split('_')[0]}}">${{t.outcome}}</span></td>
      <td class="${{ptsCls}}">${{ptsStr}}</td>
      <td class="conf-${{t.conf_level}}">${{t.confidence||'–'}}${{t.confidence?' / '+t.conf_level:''}}</td>
      <td style="color:var(--text2)">${{t.mtf||'–'}}</td>
      <td style="color:var(--text2)">${{t.tf5||'–'}}</td>
    </tr>`;
  }}).join('');
}}

function filterTrades(f){{ renderTable(f); }}
renderTable('ALL');
</script>
</body>
</html>"""
    return html


# =========================
# MAIN
# =========================

def run(open_browser: bool = False) -> Path:
    """Generate the dashboard HTML and return the output path."""
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    nojekyll = DOCS_DIR / ".nojekyll"
    if not nojekyll.exists():
        nojekyll.touch()

    rows       = _load_xlsx_files()
    validation = _load_validation_metrics()
    active     = _load_active_trade()
    kpis       = compute_kpis(rows)

    html = build_html(kpis, validation, active)
    OUTPUT_HTML.write_text(html, encoding="utf-8")

    logger.info(f"[DASH] Dashboard written: {OUTPUT_HTML} ({len(rows)} trades)")
    print(f"[DASH] Dashboard generated: {OUTPUT_HTML}")
    print(f"[DASH] Trades: {kpis['total_trades']} closed | P&L: {kpis['total_pts']} pts | Win rate: {kpis['win_rate']}%")

    if open_browser:
        webbrowser.open(OUTPUT_HTML.as_uri())
        print("[DASH] Opening in browser...")

    return OUTPUT_HTML


if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
    parser = argparse.ArgumentParser(description="TradingBot dashboard generator")
    parser.add_argument("--open", action="store_true", help="Open in browser after generating")
    args = parser.parse_args()
    run(open_browser=args.open)
