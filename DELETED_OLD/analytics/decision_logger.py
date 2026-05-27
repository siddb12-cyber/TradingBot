"""
analytics/decision_logger.py
============================
Decision Journal Logger — records every scan cycle decision to a
formatted Excel workbook, one file per trading day.

Each row is one scan cycle with full context:
  - Time and scan counter
  - OCR values from all three timeframes (5m, 15m, 1h)
  - Per-TF direction and alignment summary
  - OI analysis (PCR, max pain, ATM bias, OI score adjustment)
  - News sentiment (VIX, US Futures %, mood, sentiment adjustment)
  - Confidence scoring breakdown (base → OI adj → sentiment adj → final)
  - Decision type and full reason string
  - Trade ID (populated only on TRADE SIGNAL decisions)

Output files:
  trade_logs/YYYY-MM-DD/decisions_YYYY-MM-DD.xlsx

Usage in ai_trading_assistant.py:
  dlogger = DecisionLogger()
  dlogger.log(
      decision   = "TRADE SIGNAL",
      reason     = "MTF 3/3 aligned, BULLISH",
      trade_id   = trade_id,
      mtf        = sig,
      oi_result  = oi_result,
      sent_result= sent_result,
      base_score = base_score,
      adjusted_score = adjusted_score,
      confidence_level = level,
      scan_number = scan_n,
  )
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.config import TRADE_LOG_DIR

# =========================
# LOGGING
# =========================

logger = logging.getLogger(__name__)

# =========================
# STYLE CONSTANTS
# =========================

# Header row styling
_HEADER_BG    = "1A237E"   # Deep navy
_HEADER_FG    = "FFFFFF"   # White text
_ALT_ROW_BG   = "F5F5F5"   # Subtle grey for alternating rows

# Decision-type background fills (hex, no #)
_DECISION_BG = {
    "TRADE SIGNAL":   "1B5E20",   # Deep green
    "ACTIVE TRADE":   "BBDEFB",   # Light blue
    "SIDEWAYS":       "FFF9C4",   # Pale yellow
    "LOW CONFIDENCE": "FFE0B2",   # Soft orange
    "BLOCKED":        "FFCDD2",   # Soft red
    "OCR ERROR":      "E1BEE7",   # Soft purple
}

# Decision-type foreground (text) colour
_DECISION_FG = {
    "TRADE SIGNAL": "FFFFFF",     # White text on dark green
}

# Column definitions: (header label, column width, number format or None)
_COLUMNS = [
    ("Time",        12,  None),
    ("Scan #",       7,  "0"),
    ("Price 5m",    11,  "0.00"),
    ("VWAP 5m",     11,  "0.00"),
    ("EMA9 5m",     11,  "0.00"),
    ("vs VWAP",     10,  "+0.00;-0.00"),
    ("vs EMA9",     10,  "+0.00;-0.00"),
    ("5m Dir",      10,  None),
    ("15m Dir",     10,  None),
    ("1h Dir",       9,  None),
    ("TF Align",    16,  None),
    ("Base Score",  11,  "0"),
    ("PCR",          8,  "0.000"),
    ("Max Pain",    10,  "0"),
    ("ATM Bias",    10,  None),
    ("OI Adj",       9,  "+0;-0"),
    ("VIX",          8,  "0.00"),
    ("ES/F %",       9,  "+0.00;-0.00"),
    ("Mood",        10,  None),
    ("Sent Adj",     9,  "+0;-0"),
    ("Final Score", 11,  "0"),
    ("Confidence",  14,  None),
    ("Decision",    18,  None),
    ("Reason",      52,  None),
    ("Trade ID",    24,  None),
]

_HEADERS    = [c[0] for c in _COLUMNS]
_COL_WIDTHS = [c[1] for c in _COLUMNS]
_COL_FMTS   = [c[2] for c in _COLUMNS]

# Thin border for cell edges
_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# =========================
# HELPERS
# =========================

def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _make_font(bold: bool = False, color: str = "000000", size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size)


def _apply_header_style(ws) -> None:
    """Apply header row formatting (row 1)."""
    header_fill = _make_fill(_HEADER_BG)
    header_font = _make_font(bold=True, color=_HEADER_FG, size=10)

    for col_idx, (label, width, _) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value       = label
        cell.fill        = header_fill
        cell.font        = header_font
        cell.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border      = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28


def _apply_data_row_style(ws, row_idx: int, decision: str) -> None:
    """Colour-code one data row based on the decision type."""
    # Background: decision-specific or alternating grey
    if decision in _DECISION_BG:
        bg = _DECISION_BG[decision]
    else:
        bg = _ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"

    fg = _DECISION_FG.get(decision, "212121")  # Near-black default text

    row_fill = _make_fill(bg)
    row_font = _make_font(color=fg, size=10)

    for col_idx in range(1, len(_COLUMNS) + 1):
        cell          = ws.cell(row=row_idx, column=col_idx)
        cell.fill     = row_fill
        cell.font     = row_font
        cell.border   = _BORDER
        cell.alignment = Alignment(
            horizontal = "center" if col_idx != len(_COLUMNS) - 1 else "left",  # Reason col left-aligned
            vertical   = "center",
            wrap_text  = col_idx == len(_COLUMNS) - 1,  # Wrap Reason column
        )
        # Apply number format if defined for this column
        fmt = _COL_FMTS[col_idx - 1]
        if fmt:
            cell.number_format = fmt

    ws.row_dimensions[row_idx].height = 18


# =========================
# DECISION LOGGER CLASS
# =========================

class DecisionLogger:
    """
    Appends one row per scan cycle to the daily decisions Excel file.

    File: trade_logs/YYYY-MM-DD/decisions_YYYY-MM-DD.xlsx

    Thread/process safety: single writer (ai_trading_assistant runs in its
    own process), so no lock is needed.  Uses atomic write (tmp → replace).
    """

    def __init__(self) -> None:
        self._scan_counter: int = 0
        logger.info("[DLOG] DecisionLogger initialised (output: trade_logs/<date>/decisions_<date>.xlsx)")

    # ------------------------------------------------------------------
    # PUBLIC API
    # ------------------------------------------------------------------

    def log(
        self,
        decision:          str,
        reason:            str            = "",
        trade_id:          str            = "",
        mtf:               Optional[Dict] = None,
        oi_result:         Optional[Dict] = None,
        sent_result:       Optional[Dict] = None,
        base_score:        int            = 0,
        adjusted_score:    int            = 0,
        confidence_level:  str            = "N/A",
        scan_number:       int            = 0,
    ) -> None:
        """
        Record one decision entry.

        Parameters
        ----------
        decision        : One of: TRADE SIGNAL / ACTIVE TRADE / SIDEWAYS /
                          LOW CONFIDENCE / BLOCKED / OCR ERROR
        reason          : Human-readable explanation (e.g. risk_check["reason"])
        trade_id        : Populated only for TRADE SIGNAL rows
        mtf             : Full mtf dict from MultiTimeframeAnalyzer.analyze()
        oi_result       : Dict from OIAnalysis.get_score_adjustment()
        sent_result     : Dict from NewsSentimentEngine.get_score_adjustment()
        base_score      : Base MTF confidence score (before OI/sentiment adj)
        adjusted_score  : Final score after all adjustments
        confidence_level: HIGH / MEDIUM / LOW / VERY HIGH
        scan_number     : Global scan counter from the main loop
        """
        self._scan_counter += 1

        try:
            row = self._build_row(
                decision         = decision,
                reason           = reason,
                trade_id         = trade_id,
                mtf              = mtf or {},
                oi_result        = oi_result or {},
                sent_result      = sent_result or {},
                base_score       = base_score,
                adjusted_score   = adjusted_score,
                confidence_level = confidence_level,
                scan_number      = scan_number or self._scan_counter,
            )

            target_file = self._get_today_file()
            self._append_row(target_file, row, decision)
            logger.debug("[DLOG] Logged '%s' → %s", decision, target_file.name)

        except Exception as exc:
            # Never let logging errors crash the main trading loop
            logger.error("[DLOG] Failed to log decision '%s': %s", decision, exc)

    # ------------------------------------------------------------------
    # INTERNALS
    # ------------------------------------------------------------------

    def _get_today_file(self) -> Path:
        """Return path to today's decision log file, creating the folder if needed."""
        today     = datetime.now().strftime("%Y-%m-%d")
        day_dir   = TRADE_LOG_DIR / today
        day_dir.mkdir(parents=True, exist_ok=True)
        return day_dir / f"decisions_{today}.xlsx"

    def _build_row(
        self,
        decision:         str,
        reason:           str,
        trade_id:         str,
        mtf:              dict,
        oi_result:        dict,
        sent_result:      dict,
        base_score:       int,
        adjusted_score:   int,
        confidence_level: str,
        scan_number:      int,
    ) -> list:
        """Extract all 25 column values from the input dicts."""
        now = datetime.now()
        tf_data = mtf.get("timeframe_data", {})

        # ---- 5m values ----
        d5m    = tf_data.get("5m", {})
        price  = d5m.get("price")
        vwap   = d5m.get("vwap")
        ema9   = d5m.get("ema9")

        # ---- Derived vs-VWAP / vs-EMA9 ----
        # Use explicit None checks — do NOT use truthiness (0.0 is falsy but valid)
        vs_vwap = round(price - vwap, 2) if (price is not None and vwap is not None) else None
        vs_ema9 = round(price - ema9, 2) if (price is not None and ema9 is not None) else None

        # ---- 15m / 1h directions ----
        dir_5m  = d5m.get("direction",                         "N/A")
        dir_15m = tf_data.get("15m", {}).get("direction",     "N/A")
        dir_1h  = tf_data.get("1h",  {}).get("direction",     "N/A")

        # ---- Alignment ----
        tf_align = mtf.get("alignment_summary", "N/A")

        # ---- OI values ----
        pcr      = oi_result.get("pcr")
        max_pain = oi_result.get("max_pain")
        atm_bias = oi_result.get("atm_bias",        "N/A")
        oi_adj   = oi_result.get("score_adjustment",  0)

        # ---- Sentiment values ----
        vix      = sent_result.get("vix")
        us_pct   = sent_result.get("us_futures_pct")
        mood     = sent_result.get("sentiment",       "N/A")
        sent_adj = sent_result.get("total_adjustment",  0)

        return [
            now.strftime("%H:%M:%S"),   # A - Time
            scan_number,                 # B - Scan #
            price,                       # C - Price 5m
            vwap,                        # D - VWAP 5m
            ema9,                        # E - EMA9 5m
            vs_vwap,                     # F - vs VWAP
            vs_ema9,                     # G - vs EMA9
            dir_5m,                      # H - 5m Dir
            dir_15m,                     # I - 15m Dir
            dir_1h,                      # J - 1h Dir
            tf_align,                    # K - TF Align
            base_score or None,          # L - Base Score
            pcr,                         # M - PCR
            max_pain,                    # N - Max Pain
            atm_bias,                    # O - ATM Bias
            oi_adj if oi_adj else None,  # P - OI Adj
            vix,                         # Q - VIX
            us_pct,                      # R - ES/F %
            mood,                        # S - Mood
            sent_adj if sent_adj else None, # T - Sent Adj
            adjusted_score or None,      # U - Final Score
            confidence_level,            # V - Confidence
            decision,                    # W - Decision
            reason[:500] if reason else "", # X - Reason (truncate safety)
            trade_id or "",             # Y - Trade ID
        ]

    def _append_row(self, file_path: Path, row_data: list, decision: str) -> None:
        """
        Append one row to the Excel file.
        If the file does not exist, create it with a formatted header first.
        Uses atomic tmp→replace write to avoid corrupt files on crash.
        """
        tmp_path = file_path.with_suffix(".tmp.xlsx")

        # --- Load or create workbook ---
        if file_path.exists():
            try:
                wb = load_workbook(file_path)
                ws = wb.active
            except Exception as exc:
                logger.warning("[DLOG] Could not open existing workbook (%s) — creating fresh.", exc)
                wb = Workbook()
                ws = wb.active
                ws.title = "Decisions"
                _apply_header_style(ws)
                ws.freeze_panes = "A2"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Decisions"
            _apply_header_style(ws)
            ws.freeze_panes = "A2"

        # --- Determine next row ---
        next_row = ws.max_row + 1

        # --- Write data ---
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=next_row, column=col_idx).value = value

        # --- Style the new row ---
        _apply_data_row_style(ws, next_row, decision)

        # --- Atomic save ---
        try:
            wb.save(str(tmp_path))
            os.replace(str(tmp_path), str(file_path))
        except Exception as exc:
            logger.error("[DLOG] Atomic save failed: %s", exc)
            # Clean up tmp file if it exists
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
